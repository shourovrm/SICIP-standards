#!/usr/bin/env python3
"""Assemble per-course structured data for the site build -> data/peer-data.json.
Sources: lab-standards/*.xlsx (equipment/trainees/space), docs/reference-lab-index.html
(sector map), competency-standards/<ORG>/*.pdf (approved date from filename)."""
import os, re, glob, json, html
import openpyxl

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
LAB = os.path.join(REPO, "lab-standards")
CS = os.path.join(REPO, "competency-standards")
IDX = os.path.join(REPO, "docs", "reference-lab-index.html")

MON = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,'sept':9,
 'oct':10,'nov':11,'dec':12,'january':1,'february':2,'march':3,'april':4,'june':6,
 'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}
MN = ['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

def parse_date(fn):
    s = re.sub(r"[_']", " ", fn)
    m = re.search(r'(\d{1,2})\s*([A-Za-z]{3,9})\.?\s*(\d{2,4})', s)
    if m and m.group(2).lower().rstrip('.') in MON:
        d, mo, y = int(m.group(1)), MON[m.group(2).lower().rstrip('.')], int(m.group(3))
        y = 2000 + y if y < 100 else y
        if 1 <= d <= 31: return f"{d:02d} {MN[mo]} {y}"
    m = re.search(r'(?<!\d)(\d{1,2})[ ./-](\d{1,2})[ ./-](\d{2,4})(?!\d)', s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        y = 2000 + y if y < 100 else y
        if 1 <= d <= 31 and 1 <= mo <= 12: return f"{d:02d} {MN[mo]} {y}"
    m = re.search(r'(?<!\d)(\d{2})(\d{2})(\d{4})(?!\d)', s)
    if m and 1 <= int(m.group(1)) <= 31 and 1 <= int(m.group(2)) <= 12:
        return f"{int(m.group(1)):02d} {MN[int(m.group(2))]} {m.group(3)}"
    m = re.search(r'(?<!\d)(\d{2})(\d{2})(\d{2})(?!\d)', s)
    if m and 1 <= int(m.group(1)) <= 31 and 1 <= int(m.group(2)) <= 12:
        return f"{int(m.group(1)):02d} {MN[int(m.group(2))]} 20{m.group(3)}"
    return None

def norm(s): return re.sub(r'[^a-z0-9]', '', s.lower())
def toks(s): return set(re.findall(r'[a-z0-9]+', s.lower())) - {
    'cs','final','m','the','of','and','for','on','in','bmet','dte','sicip','develop',
    'review','revised','version','v2','done'}

# sector map from reference index: (org_slug_norm, course_norm) -> sector
sectors = {}
h = open(IDX).read()
for sec in re.findall(r'<section class="org" data-org="([^"]+)">(.*?)</section>', h, re.S):
    org = norm(html.unescape(sec[0]))   # unescape so 'isc-t&amp;h' -> 'iscth'
    for li in re.findall(r'<span class="c-name">(.*?)</span><span class="c-full">(.*?)</span>', sec[1], re.S):
        cn = html.unescape(re.sub(r'<[^>]+>', '', li[0]))
        sect = None
        m = re.search(r'Sector:\s*(.+)', html.unescape(li[1]))
        if m: sect = m.group(1).strip()
        sectors[(org, norm(cn))] = sect

def match_pdf(course_name, pdfs):
    ct = toks(course_name)
    best, bs = None, 0
    for p in pdfs:
        pt = toks(os.path.splitext(p)[0])
        sc = len(ct & pt)
        if sc > bs: bs, best = sc, p
    return best if bs >= 1 else None

courses, mismatches = [], []
for wb_path in sorted(glob.glob(os.path.join(LAB, "*-lab-standard.xlsx"))):
    slug = os.path.basename(wb_path).replace("-lab-standard.xlsx", "")
    if slug.startswith("Weaving-and-Finishing"):   # standalone dup of BJMA Weaving
        mismatches.append("Weaving-and-Finishing-Jute-lab-standard.xlsx is a standalone duplicate of BJMA's 'Weaving and Finishing' sheet — excluded from courses.")
        continue
    org_dir = os.path.join(CS, slug)
    pdfs = [os.path.basename(p) for p in glob.glob(os.path.join(org_dir, "*.pdf"))] if os.path.isdir(org_dir) else []
    wb = openpyxl.load_workbook(wb_path, data_only=False)
    # 1) read all sheets for this org
    recs = []
    for ws in wb.worksheets:
        equip, r = [], 16
        while ws.cell(r, 1).value not in ("Sum", None):
            name = ws.cell(r, 2).value
            if name is None: break
            equip.append({"name": name, "required": ws.cell(r, 3).value, "weight": ws.cell(r, 5).value})
            r += 1
        recs.append({"course_name": ws["B3"].value, "sheet": ws.title,
                     "trainees": ws["B4"].value, "space": ws["C8"].value, "equipment": equip, "pdf": None})
    # 2) global best-first assignment of pdfs to sheets (avoids greedy Construction/Industry collisions)
    pairs = sorted(((len(toks(rec["course_name"] or rec["sheet"]) & toks(os.path.splitext(p)[0])), i, p)
                    for i, rec in enumerate(recs) for p in pdfs), reverse=True)
    taken_sheet, taken_pdf = set(), set()
    for score, i, p in pairs:
        if score < 1 or i in taken_sheet or p in taken_pdf: continue
        recs[i]["pdf"] = p; taken_sheet.add(i); taken_pdf.add(p)
    # 3) finalize
    for rec in recs:
        sector = sectors.get((norm(slug), norm(rec["course_name"] or "")))
        pdf = rec["pdf"]; approved = parse_date(pdf) if pdf else None
        courses.append({
            "org": slug, "org_slug": slug, "course_name": rec["course_name"], "sheet": rec["sheet"],
            "sector": sector, "trainees": rec["trainees"], "space": rec["space"],
            "approved": approved, "approved_source": ("filename" if approved else None),
            "equipment": rec["equipment"],
            "cs_pdf": (f"competency-standards/{slug}/{pdf}" if pdf else None),
        })

nodate = [f"{c['org']} / {c['course_name']}" for c in courses if not c["approved"]]
nopdf = [f"{c['org']} / {c['course_name']}" for c in courses if not c["cs_pdf"]]
nosector = [f"{c['org']} / {c['course_name']}" for c in courses if not c["sector"]]
mismatches += [
    "BPI 'Electrical Works V2' duplicate: already REMOVED from BPI workbook + indexes (BPI now 3 sheets, total 162). No action needed.",
    f"{len(nodate)} courses have no date parsed from the CS filename (need cover-page date): " + "; ".join(nodate),
    f"{len(nosector)} courses have null sector (all Beautification): " + "; ".join(nosector),
    (f"{len(nopdf)} courses did not auto-match a CS PDF by filename tokens (verify manually): " + "; ".join(nopdf)) if nopdf else "All courses matched a CS PDF.",
]

out = {
    "schema": {
        "courses": "list of course objects",
        "course": "org, org_slug, course_name, sheet, sector|null, trainees, space, "
                   "approved ('DD Mon YYYY'|null), approved_source, "
                   "equipment[{name, required(int|str), weight(int)}], cs_pdf(repo-relative|null)",
    },
    "counts": {"courses": len(courses), "with_date": len(courses) - len(nodate),
               "with_pdf": len(courses) - len(nopdf), "with_sector": len(courses) - len(nosector)},
    "known_mismatches": mismatches,
    "courses": courses,
}
os.makedirs(os.path.join(REPO, "data"), exist_ok=True)
json.dump(out, open(os.path.join(REPO, "data", "peer-data.json"), "w"), indent=1, ensure_ascii=False)
print(f"wrote data/peer-data.json: {len(courses)} courses | dates {out['counts']['with_date']} | pdf {out['counts']['with_pdf']} | sector {out['counts']['with_sector']}")
print("no-date:", len(nodate), "| no-pdf:", len(nopdf), "| no-sector:", len(nosector))
