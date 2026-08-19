#!/usr/bin/env python3
"""Build SICIP lab-standard xlsx sheets from structured course data.
build_sheet(ws, course) writes one styled sheet matching the Beautification template.
main(json_path, out_dir) builds a full workbook from {outfile, courses:[...]}.
Course: {sheet, course_name, trainees, space, equipment:[[name, qty, weight], ...]}.
qty may be a string; leading integer is used and the original kept in Remarks.
"""
import sys, os, json, re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="FFD9D9D9")
F11 = lambda b=False: Font(name="Arial", size=11, bold=b)
F14 = Font(name="Arial", size=14, bold=True)
CEN = lambda wrap=False: Alignment(horizontal="center", vertical="center", wrap_text=wrap)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

BOILER = (
    "•The institute shall not use the same facilities for any other "
    "projects/organizations offering a similar course.\n"
    "•The institute must provide sufficient evidence to prove ownership of "
    "the proposed training equipment.\n"
    "The list denotes the minimum training equipment and facility required to "
    "effectively conduct training for a specific course. Additionally, the "
    "institute must ensure that all other necessary training tools, equipment, "
    "and furniture are available to meet the requirement of competency standards "
    "(CS) provided by SICIP.\n"
    "For the operation of training course on {course}, the institute must ensure "
    "the availability of at least 80% of the major training equipment and "
    "training facilities (according to the CS) to be eligible for SICIP training "
    "delivery. If the score is below 80%, the remaining equipment and facilities "
    "need to be installed before the commencement of the training.\n"
    "The institute will also provide all other hand tools and power tools as per "
    "CS for {n} trainees. Also, they will arrange adequate seating arrangement "
    "and classroom setup for the {n} trainees."
)
INVALID = re.compile(r'[:\\/?*\[\]]')

def safe_sheet(name, used):
    name = INVALID.sub(" ", str(name)).strip()[:31] or "Sheet"
    base, i = name, 1
    while name.lower() in used:
        suf = f" {i}"; name = base[:31 - len(suf)] + suf; i += 1
    used.add(name.lower()); return name

def set_cell(ws, coord, val, font=None, fill=None, align=None, border=True):
    c = ws[coord]; c.value = val
    if font: c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = BORDER
    return c

def build_sheet(ws, co):
    course = co["course_name"]; n = co.get("trainees", 25); equip = co["equipment"]
    ws.column_dimensions["A"].width = 15.17
    ws.column_dimensions["B"].width = 33.13
    for col in "CDEFG": ws.column_dimensions[col].width = 13.72

    ws.merge_cells("A1:G1")
    set_cell(ws, "A1", "Course-wise Training Infrastructure and Facilities",
             font=F14, align=Alignment(horizontal="center", vertical="bottom"), border=False)
    ws.row_dimensions[1].height = 17.35
    set_cell(ws, "A3", "Course Name:", font=F11(True), align=CEN(True))
    set_cell(ws, "B3", course, font=F11(True), align=CEN(True))
    set_cell(ws, "A4", "Number of Trainees:", font=F11(True), align=CEN(True))
    set_cell(ws, "B4", n, font=F11(True), align=CEN(True))
    ws.row_dimensions[4].height = 26.85
    set_cell(ws, "A6", "Course-wise Training Space (Theoretical Classroom, "
             "Workshop/ Lab/ Classroom cum Workshop)", font=F11(True), border=False)
    set_cell(ws, "B7", "Course Name", font=F11(True), align=CEN(True))
    ws.merge_cells("C7:G7")
    set_cell(ws, "C7", f"SICIP required space for {n} trainees", font=F11(True), align=CEN(True))
    set_cell(ws, "B8", course, font=F11(True), align=CEN(True))
    ws.merge_cells("C8:G8")
    set_cell(ws, "C8", co.get("space", ""), font=F11(True),
             align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[8].height = 52.2
    set_cell(ws, "A11", "Major Training Equipment and Training Facilities", font=F11(True), border=False)
    ws.merge_cells("A13:G13")
    set_cell(ws, "A13", BOILER.format(course=course, n=n), font=F11(False), align=LEFT, border=False)
    ws.row_dimensions[13].height = 115.65

    hdr = ["S.N.", "Major Equipment and Training facilities", "Required facilities",
           "Available facilities", "Weights\n(out of 10)", "Weighted scores", "Remarks"]
    for col, txt in zip("ABCDEFG", hdr):
        set_cell(ws, f"{col}15", txt, font=F11(True), fill=HDR_FILL, align=CEN(True))
    ws.row_dimensions[15].height = 26.85

    r0 = 16
    for i, item in enumerate(equip):
        r = r0 + i
        name, raw_qty = item[0], item[1]
        weight = item[2] if len(item) > 2 else None
        remark = None
        if isinstance(raw_qty, bool): qty = None
        elif isinstance(raw_qty, int): qty = raw_qty
        else:
            m = re.match(r"\s*(\d+)", str(raw_qty))
            qty = int(m.group(1)) if m else None
            remark = str(raw_qty).strip() or None
        set_cell(ws, f"A{r}", i + 1, font=F11(), align=CEN())
        set_cell(ws, f"B{r}", name, font=F11(), align=Alignment(horizontal="left", vertical="top", wrap_text=True))
        set_cell(ws, f"C{r}", qty, font=F11(), align=CEN(True))
        set_cell(ws, f"D{r}", None, font=F11(), align=CEN(True))
        set_cell(ws, f"E{r}", weight, font=F11(), align=CEN(True))
        set_cell(ws, f"F{r}", f"=IFERROR(MIN(C{r},D{r})*E{r}/C{r},0)", font=F11(), align=CEN(True))
        set_cell(ws, f"G{r}", remark, font=F11(), align=CEN(True))
        ws.row_dimensions[r].height = 14.15

    last = r0 + len(equip) - 1
    sum_r, score_r, spacer_r, pts_r = last + 1, last + 2, last + 3, last + 4
    set_cell(ws, f"A{sum_r}", "Sum", font=F11(True), align=CEN())
    for col in "BCD": set_cell(ws, f"{col}{sum_r}", None, font=F11(True))
    ws.merge_cells(f"A{sum_r}:D{sum_r}")
    set_cell(ws, f"E{sum_r}", f"=SUM(E{r0}:E{last})", font=F11(True), align=CEN())
    set_cell(ws, f"F{sum_r}", f"=SUM(F{r0}:F{last})", font=F11(True), align=CEN())
    set_cell(ws, f"G{sum_r}", None, font=F11(True))
    set_cell(ws, f"A{score_r}", "Score out of 100", font=F11(True), align=CEN())
    for col in "BCDE": set_cell(ws, f"{col}{score_r}", None, font=F11(True))
    ws.merge_cells(f"A{score_r}:E{score_r}")
    set_cell(ws, f"F{score_r}", f"=F{sum_r}/E{sum_r}*100", font=F11(True), align=CEN())
    set_cell(ws, f"G{score_r}", None, font=F11(True))
    set_cell(ws, f"A{spacer_r}", None, border=False)
    ws.merge_cells(f"A{spacer_r}:G{spacer_r}")
    ws.row_dimensions[spacer_r].height = 9.75
    set_cell(ws, f"A{pts_r}", "Total achieved points out of 30  ", font=F11(True), align=CEN())
    for col in "BCDE": set_cell(ws, f"{col}{pts_r}", None, font=F11(True))
    ws.merge_cells(f"A{pts_r}:E{pts_r}")
    set_cell(ws, f"F{pts_r}", f"=F{score_r}*30/100", font=F11(True), align=CEN())
    set_cell(ws, f"G{pts_r}", "Points", font=F11(True), align=CEN())

def main(json_path, out_dir):
    d = json.load(open(json_path))
    wb = openpyxl.Workbook(); wb.remove(wb.active); used = set()
    for co in d["courses"]:
        build_sheet(wb.create_sheet(safe_sheet(co["sheet"], used)), co)
    out = os.path.join(out_dir, d["outfile"]); wb.save(out)
    print(f"wrote {out}  ({len(d['courses'])} sheets)")

if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
