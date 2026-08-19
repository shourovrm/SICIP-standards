#!/usr/bin/env python3
"""Build data/data.json from lab-standards/*.xlsx + docs/reference-lab-index.html.
Uses data/peer-data.json (another session's extraction) for approved-date and
cs_pdf matching (already fuzzy-resolved there); re-derives sector, course list,
and equipment/boilerplate fresh from the authoritative sources per CLAUDE.md.
"""
import html
import json
import re
import subprocess
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
LAB_DIR = ROOT / "lab-standards"
CS_DIR = ROOT / "competency-standards"
INDEX_HTML = ROOT / "docs" / "reference-lab-index.html"
PEER = ROOT / "data" / "peer-data.json"
OUT = ROOT / "data" / "data.json"

MONTHS = {m.lower(): i for i, m in enumerate(
    "Jan Feb Mar Apr May Jun Jul Aug Sep Oct Nov Dec".split(), 1)}
MONTHS_FULL = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11,
    "december": 12,
}


def norm(s):
    s = s.lower().replace("&", "and")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# ---------- 1. parse reference-lab-index.html (authoritative course list) ----------

def parse_index():
    raw = INDEX_HTML.read_text(encoding="utf-8")
    orgs = []
    for sec in re.finditer(r'<section class="org".*?</section>', raw, re.S):
        block = sec.group()
        org_name = html.unescape(re.search(r'class="org-name"[^>]*>([^<]*)<', block).group(1))
        href = re.search(r'class="org-name" href="\./([^"]+)-lab-standard\.xlsx"', block).group(1)
        org_slug = re.sub(r"%26", "&", href)
        courses = []
        for li in re.finditer(r'<span class="c-name">([^<]*)</span><span class="c-full">([^<]*)</span>', block):
            cname = html.unescape(li.group(1))
            full = html.unescape(li.group(2))
            sm = re.match(r"Sector:\s*(.+)", full)
            sector = sm.group(1).strip() if sm else None
            courses.append((cname, sector))
        orgs.append((org_name, org_slug, courses))
    return orgs


# ---------- 2. xlsx sheet lookup: find sheet whose B3 matches course name ----------

def find_sheet(wb, course_name):
    target = norm(course_name)
    candidates = []
    for name in wb.sheetnames:
        ws = wb[name]
        b3 = ws["B3"].value or ""
        if norm(b3) == target:
            candidates.append(name)
    if candidates:
        # BPI-style dup: prefer sheet with higher revision marker in name
        if len(candidates) > 1:
            candidates.sort(key=lambda n: ("v2" in n.lower() or "revised" in n.lower()))
            return candidates[-1]
        return candidates[0]
    # fallback: fuzzy sheet-name match
    for name in wb.sheetnames:
        if norm(name) == target or target in norm(name) or norm(name) in target:
            return name
    return None


def read_sheet(ws):
    course_name = ws["B3"].value
    trainees = ws["B4"].value
    space = ws["C8"].value
    boilerplate = ws["A13"].value
    equip = []
    r = 16
    while True:
        a = ws.cell(r, 1).value
        if a is None and ws.cell(r, 2).value is None:
            break
        if isinstance(a, str) and a.strip().lower() == "sum":
            break
        name = ws.cell(r, 2).value
        if name:
            equip.append({
                "name": str(name).strip(),
                "required": ws.cell(r, 3).value,
                "weight": ws.cell(r, 5).value,
                "remark": ws.cell(r, 7).value or None,
            })
        r += 1
    return course_name, trainees, space, boilerplate, equip


# ---------- 2b. cs_pdf fallback fuzzy match (when peer-data has none) ----------

def fuzzy_match_pdf(org_slug, course_name):
    d = CS_DIR / org_slug
    if not d.is_dir():
        return None
    target = norm(course_name)
    target_sq = target.replace(" ", "")
    target_tokens = set(target.split())
    sig_tokens = [t for t in target_tokens if len(t) >= 4]
    best, best_ratio = None, 0.0
    for pdf in d.glob("*.pdf"):
        cand = norm(pdf.stem)
        cand_sq = cand.replace(" ", "")
        if target in cand or cand in target or target_sq in cand_sq or cand_sq in target_sq:
            return f"competency-standards/{org_slug}/{pdf.name}"
        hits = sum(1 for t in sig_tokens if t in cand_sq)
        ratio = hits / len(sig_tokens) if sig_tokens else 0
        if ratio > best_ratio:
            best, best_ratio = pdf, ratio
    if best and best_ratio >= 0.6:
        return f"competency-standards/{org_slug}/{best.name}"
    return None


# ---------- 3. approved-date fallback: pdftotext cover page ----------

DATE_RES = [
    re.compile(r"(\d{1,2})\s+(" + "|".join(MONTHS) + r")[a-z]*\s+'?(\d{2,4})", re.I),
    re.compile(r"(\d{1,2})\s+(" + "|".join(MONTHS_FULL) + r")\s+'?(\d{2,4})", re.I),
]


def date_from_text(text):
    for rx in DATE_RES:
        m = rx.search(text)
        if m:
            d, mo, y = m.groups()
            mo_l = mo.lower()
            month = MONTHS.get(mo_l[:3], MONTHS_FULL.get(mo_l))
            if not month:
                continue
            y = int(y)
            if y < 100:
                y += 2000
            try:
                return f"{int(d):02d} {list(MONTHS)[month-1].capitalize()} {y}"
            except Exception:
                continue
    return None


def extract_pdf_date(pdf_path):
    try:
        out = subprocess.run(["pdftotext", "-l", "1", str(pdf_path), "-"],
                              capture_output=True, text=True, timeout=30).stdout
    except Exception:
        return None
    return date_from_text(out)


def main():
    peer = json.loads(PEER.read_text(encoding="utf-8"))
    peer_by_key = {(c["org_slug"], norm(c["course_name"])): c for c in peer["courses"]}

    orgs = parse_index()
    entries = []
    unmatched_sheets = []
    no_peer = []

    for org_name, org_slug, courses in orgs:
        xlsx_path = LAB_DIR / f"{org_slug}-lab-standard.xlsx"
        if not xlsx_path.exists():
            print(f"MISSING WORKBOOK: {xlsx_path}")
            continue
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        for course_name, sector in courses:
            sheet_name = find_sheet(wb, course_name)
            if sheet_name is None:
                unmatched_sheets.append((org_slug, course_name))
                continue
            ws = wb[sheet_name]
            b3, trainees, space, boilerplate, equip = read_sheet(ws)

            key = (org_slug, norm(course_name))
            pc = peer_by_key.get(key)
            if pc is None:
                no_peer.append((org_slug, course_name))
                approved, cs_pdf = None, None
            else:
                approved, cs_pdf = pc.get("approved"), pc.get("cs_pdf")

            if not cs_pdf:
                cs_pdf = fuzzy_match_pdf(org_slug, course_name)

            if approved is None and cs_pdf:
                approved = date_from_text(Path(cs_pdf).stem)
            if approved is None and cs_pdf:
                pdf_path = ROOT / cs_pdf
                if pdf_path.exists():
                    approved = extract_pdf_date(pdf_path)

            entries.append({
                "org": org_name,
                "org_slug": org_slug,
                "course_name": course_name,
                "sheet": sheet_name,
                "sector": sector,
                "trainees": int(trainees) if trainees is not None else None,
                "space": str(space).strip() if space else "",
                "boilerplate": str(boilerplate).strip() if boilerplate else "",
                "approved": approved,
                "equipment": equip,
                "xlsx": f"lab-standards/{xlsx_path.name}",
                "cs_pdf": cs_pdf,
            })

    OUT.write_text(json.dumps(entries, ensure_ascii=False, indent=1), encoding="utf-8")

    # ---------------- validation ----------------
    print(f"\n=== VALIDATION REPORT ===")
    print(f"total entries: {len(entries)} (expect 162)")
    assert len(entries) == 162, f"expected 162 entries, got {len(entries)}"

    seen = set()
    dupes = []
    for e in entries:
        k = (e["org_slug"], e["course_name"])
        if k in seen:
            dupes.append(k)
        seen.add(k)
    print(f"duplicate (org_slug, course_name): {dupes}")
    assert not dupes

    bad_required, bad_weight = [], []
    for e in entries:
        assert e["equipment"], f"empty equipment: {e['org_slug']}/{e['course_name']}"
        for row in e["equipment"]:
            if not row["required"] or row["required"] <= 0:
                bad_required.append((e["org_slug"], e["course_name"], row["name"]))
            if not row["weight"] or row["weight"] <= 0:
                bad_weight.append((e["org_slug"], e["course_name"], row["name"]))
        assert sum(r["weight"] or 0 for r in e["equipment"]) > 0
        assert isinstance(e["trainees"], int), (e["org_slug"], e["course_name"], e["trainees"])
        xp = ROOT / e["xlsx"]
        assert xp.exists(), f"missing xlsx: {xp}"
        if e["cs_pdf"]:
            assert (ROOT / e["cs_pdf"]).exists(), f"missing pdf: {e['cs_pdf']}"

    print(f"\nequipment rows with required<=0/null ({len(bad_required)}) [source-data anomaly, not a script bug]:")
    for x in bad_required:
        print(" ", x)
    print(f"equipment rows with weight<=0/null ({len(bad_weight)}):")
    for x in bad_weight:
        print(" ", x)

    no_pdf = [(e["org_slug"], e["course_name"]) for e in entries if not e["cs_pdf"]]
    no_date = [(e["org_slug"], e["course_name"]) for e in entries if not e["approved"]]
    no_sector = [(e["org_slug"], e["course_name"]) for e in entries if not e["sector"]]
    no_80 = [(e["org_slug"], e["course_name"]) for e in entries if "80%" not in e["boilerplate"]]

    print(f"\nentries with cs_pdf=null ({len(no_pdf)}):")
    for x in no_pdf:
        print(" ", x)
    print(f"\nentries with approved=null ({len(no_date)}):")
    for x in no_date:
        print(" ", x)
    print(f"\nentries with sector=null ({len(no_sector)}) [expect 5, all Beautification]:")
    for x in no_sector:
        print(" ", x)
    print(f"\nentries missing '80%' in boilerplate ({len(no_80)}):")
    for x in no_80:
        print(" ", x)

    print(f"\nunmatched sheets (course in index, no xlsx sheet found): {unmatched_sheets}")
    print(f"no peer-data match (approved/cs_pdf left null/needs pdftotext): {no_peer}")


if __name__ == "__main__":
    main()
