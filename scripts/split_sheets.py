#!/usr/bin/env python3
"""Split each org workbook into per-course single-sheet xlsx files.

Output: data/xlsx/<slug>.xlsx where <slug> matches web/js/app.js cSlug().
Template formatting is preserved (openpyxl round-trip of openpyxl-built files);
fullCalcOnLoad makes Excel/LO recalc the F-column formulas on open.
"""
import json, re, sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent

def slug(s):
    s = s.lower().replace("&", "and")
    return re.sub(r"[^a-z0-9]+", "-", s).strip("-")

def cslug(e):
    org = re.sub(r"[^a-z0-9]+", "-", e["org_slug"].lower()).strip("-")
    return org + "--" + slug(e["course_name"])

def main():
    data = json.loads((ROOT / "data/data.json").read_text())
    out = ROOT / "data/xlsx"
    out.mkdir(exist_ok=True)
    for e in data:
        wb = openpyxl.load_workbook(ROOT / e["xlsx"])
        for name in [n for n in wb.sheetnames if n != e["sheet"]]:
            del wb[name]
        assert wb.sheetnames == [e["sheet"]], (e["xlsx"], e["sheet"])
        wb.calculation.fullCalcOnLoad = True
        wb.save(out / (cslug(e) + ".xlsx"))
    print(f"wrote {len(data)} files to {out}")

if __name__ == "__main__":
    sys.exit(main())
